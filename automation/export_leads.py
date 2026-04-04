# automation/export_leads.py
import pandas as pd
import os
from datetime import datetime


def export_leads_to_csv():
    """
    Export leads to a timestamped CSV file.
    Also saves a master file with all leads combined.
    """
    leads_file = "data/final_leads.csv"

    if not os.path.exists(leads_file):
        print("❌ No final_leads.csv found. Run the pipeline first.")
        return None

    df = pd.read_csv(leads_file)
    if df.empty:
        print("❌ No leads found.")
        return None

    # Create exports directory if it doesn't exist
    export_dir = "data/exports"
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)

    # Create timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_file = os.path.join(export_dir, f"leads_{timestamp}.csv")

    # Save timestamped export
    df.to_csv(export_file, index=False)
    print(f"\n✅ Leads exported to: {export_file}")
    print(f"   Total leads: {len(df)}")

    # Also save to master file (appends all exports)
    master_file = os.path.join(export_dir, "all_leads_master.csv")
    if os.path.exists(master_file):
        # Append to existing master file
        existing_df = pd.read_csv(master_file)
        combined_df = pd.concat([existing_df, df], ignore_index=True)
        combined_df.to_csv(master_file, index=False)
        print(f"   Updated master file: {master_file} (total: {len(combined_df)} leads)")
    else:
        # Create new master file
        df.to_csv(master_file, index=False)
        print(f"   Created master file: {master_file}")

    return export_file


def print_leads_summary():
    """Print formatted leads summary to console."""
    leads_file = "data/final_leads.csv"

    if not os.path.exists(leads_file):
        print("❌ No final_leads.csv found.")
        return

    df = pd.read_csv(leads_file)
    if df.empty:
        print("❌ No leads found.")
        return

    print("\n" + "=" * 70)
    print("📞 LEADS WITH PHONE NUMBERS - SUMMARY")
    print("=" * 70)

    # Summary statistics
    print(f"\n📊 Total leads: {len(df)}")

    if 'Price' in df.columns:
        # Count properties by price range
        price_col = df['Price'].astype(str)
        crores = price_col[price_col.str.contains('Crore', na=False)]
        lacs = price_col[price_col.str.contains('Lac', na=False)]
        print(f"   Properties > 1 Crore: {len(crores)}")
        print(f"   Properties < 1 Crore: {len(lacs)}")

    if 'Location' in df.columns:
        print(f"   Unique locations: {df['Location'].nunique()}")

    if 'Seller' in df.columns:
        print(f"   Unique sellers: {df['Seller'].nunique()}")

    if 'Furnished' in df.columns:
        furnished = df[df['Furnished'].str.contains('Furnished', na=False)]
        print(f"   Furnished properties: {len(furnished)}")

    if 'Bedrooms' in df.columns:
        # Extract bedroom numbers
        beds = df['Bedrooms'].dropna()
        print(f"   Properties with bedrooms: {len(beds)}")

    # List all leads
    print("\n" + "-" * 70)
    print("📋 LEAD DETAILS")
    print("-" * 70)

    for idx, row in df.iterrows():
        print(f"\n{idx + 1}. {row.get('Title', 'N/A')[:70]}")
        print(f"   💰 Price: {row.get('Price', 'N/A')}")
        print(f"   📍 Location: {row.get('Location', 'N/A')}")
        print(f"   📞 Phone: {row.get('Phone', 'N/A')}")
        print(f"   🏠 Area: {row.get('Area', 'N/A')}")
        print(f"   🛏️ Bedrooms: {row.get('Bedrooms', 'N/A')}")
        print(f"   🚿 Bathrooms: {row.get('Bathrooms', 'N/A')}")
        print(f"   🛋️ Furnished: {row.get('Furnished', 'N/A')}")
        print(f"   👤 Seller: {row.get('Seller', 'N/A')}")
        print(f"   📅 Posted: {row.get('Posted', 'N/A')}")
        if 'ImageURL' in row and pd.notna(row.get('ImageURL')):
            print(f"   🖼️ Image: {row.get('ImageURL', 'N/A')[:80]}...")
        print(f"   🔗 Link: {row.get('Link', 'N/A')[:80]}...")

    print("\n" + "=" * 70)
    print(f"✅ Total: {len(df)} leads exported")
    print("=" * 70)


def export_to_excel():
    """
    Export leads to Excel format with formatting (requires openpyxl).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        leads_file = "data/final_leads.csv"

        if not os.path.exists(leads_file):
            print("❌ No final_leads.csv found.")
            return None

        df = pd.read_csv(leads_file)
        if df.empty:
            print("❌ No leads found.")
            return None

        # Create exports directory
        export_dir = "data/exports"
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(export_dir, f"leads_{timestamp}.xlsx")

        # Write to Excel with formatting
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leads', index=False)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Leads']

            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Format header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            worksheet.freeze_panes = 'A2'

        print(f"\n✅ Excel file exported to: {excel_file}")
        return excel_file

    except ImportError:
        print("⚠️ openpyxl not installed. Install with: pip install openpyxl")
        return None
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return None


def generate_html_report():
    """
    Generate an HTML report file for easy viewing in browser with images.
    """
    leads_file = "data/final_leads.csv"

    if not os.path.exists(leads_file):
        print("❌ No final_leads.csv found.")
        return None

    df = pd.read_csv(leads_file)
    if df.empty:
        print("❌ No leads found.")
        return None

    export_dir = "data/exports"
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_file = os.path.join(export_dir, f"leads_report_{timestamp}.html")

    # Calculate summary statistics
    total_leads = len(df)
    if 'Price' in df.columns:
        price_col = df['Price'].astype(str)
        crores = len(price_col[price_col.str.contains('Crore', na=False)])
        lacs = len(price_col[price_col.str.contains('Lac', na=False)])
    else:
        crores = lacs = 0

    if 'Location' in df.columns:
        unique_locations = df['Location'].nunique()
    else:
        unique_locations = 0

    if 'Furnished' in df.columns:
        furnished_count = len(df[df['Furnished'].str.contains('Furnished', na=False)])
    else:
        furnished_count = 0

    # Create HTML content
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Real Estate Leads Report - {timestamp}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}
        .header {{
            background: white;
            border-radius: 15px;
            padding: 25px;
            margin-bottom: 25px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            font-size: 28px;
            margin-bottom: 10px;
        }}
        .timestamp {{
            color: #7f8c8d;
            font-size: 14px;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .summary-card {{
            background: white;
            border-radius: 12px;
            padding: 20px;
            text-align: center;
            box-shadow: 0 3px 10px rgba(0,0,0,0.1);
            transition: transform 0.3s ease;
        }}
        .summary-card:hover {{
            transform: translateY(-5px);
        }}
        .summary-card .number {{
            font-size: 36px;
            font-weight: bold;
            color: #3498db;
        }}
        .summary-card .label {{
            color: #7f8c8d;
            font-size: 14px;
            margin-top: 5px;
        }}
        .leads-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(380px, 1fr));
            gap: 20px;
        }}
        .lead-card {{
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 3px 10px rgba(0,0,0,0.1);
            transition: transform 0.3s ease, box-shadow 0.3s ease;
        }}
        .lead-card:hover {{
            transform: translateY(-3px);
            box-shadow: 0 8px 25px rgba(0,0,0,0.15);
        }}
        .lead-image {{
            width: 100%;
            height: 200px;
            overflow: hidden;
            background: #f5f5f5;
        }}
        .lead-image img {{
            width: 100%;
            height: 100%;
            object-fit: cover;
            transition: transform 0.3s ease;
        }}
        .lead-card:hover .lead-image img {{
            transform: scale(1.05);
        }}
        .no-image {{
            display: flex;
            align-items: center;
            justify-content: center;
            height: 100%;
            color: #bdc3c7;
            font-size: 14px;
        }}
        .lead-content {{
            padding: 15px;
        }}
        .lead-title {{
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 10px;
            line-height: 1.4;
            height: 44px;
            overflow: hidden;
        }}
        .lead-detail {{
            margin: 8px 0;
            font-size: 13px;
            color: #555;
        }}
        .lead-label {{
            font-weight: 600;
            color: #7f8c8d;
            display: inline-block;
            width: 70px;
        }}
        .phone {{
            color: #27ae60;
            font-weight: bold;
            font-size: 14px;
        }}
        .price {{
            color: #e74c3c;
            font-weight: bold;
            font-size: 16px;
        }}
        .seller {{
            color: #8e44ad;
        }}
        .link {{
            margin-top: 10px;
            padding-top: 10px;
            border-top: 1px solid #ecf0f1;
        }}
        .link a {{
            color: #3498db;
            text-decoration: none;
            font-size: 12px;
            word-break: break-all;
        }}
        .link a:hover {{
            text-decoration: underline;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            padding: 20px;
            color: white;
            font-size: 12px;
        }}
        @media (max-width: 768px) {{
            .leads-grid {{
                grid-template-columns: 1fr;
            }}
            .lead-title {{
                height: auto;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🏠 Real Estate Lead Engine Report</h1>
            <div class="timestamp">Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>

        <div class="summary">
            <div class="summary-card">
                <div class="number">{total_leads}</div>
                <div class="label">Total Leads</div>
            </div>
            <div class="summary-card">
                <div class="number">{crores}</div>
                <div class="label">Properties > 1 Crore</div>
            </div>
            <div class="summary-card">
                <div class="number">{lacs}</div>
                <div class="label">Properties < 1 Crore</div>
            </div>
            <div class="summary-card">
                <div class="number">{unique_locations}</div>
                <div class="label">Unique Locations</div>
            </div>
            <div class="summary-card">
                <div class="number">{furnished_count}</div>
                <div class="label">Furnished Properties</div>
            </div>
        </div>

        <div class="leads-grid">
"""

    for idx, row in df.iterrows():
        # Get image URL
        image_html = ""
        if 'ImageURL' in row and pd.notna(row.get('ImageURL')) and row.get('ImageURL'):
            image_html = f'<img src="{row["ImageURL"]}" alt="Property Image" loading="lazy">'
        else:
            image_html = '<div class="no-image">📷 No Image Available</div>'

        # Get values with defaults
        title = str(row.get('Title', 'N/A'))[:100]
        price = row.get('Price', 'N/A')
        location = row.get('Location', 'N/A')
        phone = row.get('Phone', 'N/A')
        area = row.get('Area', 'N/A')
        bedrooms = row.get('Bedrooms', 'N/A')
        bathrooms = row.get('Bathrooms', 'N/A')
        furnished = row.get('Furnished', 'N/A')
        seller = row.get('Seller', 'N/A')
        posted = row.get('Posted', 'N/A')
        link = row.get('Link', '#')

        html_content += f"""
            <div class="lead-card">
                <div class="lead-image">
                    {image_html}
                </div>
                <div class="lead-content">
                    <div class="lead-title">{idx + 1}. {title}</div>
                    <div class="lead-detail"><span class="lead-label">💰 Price:</span> <span class="price">{price}</span></div>
                    <div class="lead-detail"><span class="lead-label">📍 Location:</span> {location}</div>
                    <div class="lead-detail"><span class="lead-label">📞 Phone:</span> <span class="phone">{phone}</span></div>
                    <div class="lead-detail"><span class="lead-label">🏠 Area:</span> {area}</div>
                    <div class="lead-detail"><span class="lead-label">🛏️ Bedrooms:</span> {bedrooms}</div>
                    <div class="lead-detail"><span class="lead-label">🚿 Bathrooms:</span> {bathrooms}</div>
                    <div class="lead-detail"><span class="lead-label">🛋️ Furnished:</span> {furnished}</div>
                    <div class="lead-detail"><span class="lead-label">👤 Seller:</span> <span class="seller">{seller}</span></div>
                    <div class="lead-detail"><span class="lead-label">📅 Posted:</span> {posted}</div>
                    <div class="link"><a href="{link}" target="_blank">🔗 View on OLX →</a></div>
                </div>
            </div>
"""

    html_content += f"""
        </div>
        <div class="footer">
            Generated by Real Estate Lead Engine | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
            Data includes {total_leads} property leads with contact information
        </div>
    </div>
</body>
</html>
"""

    with open(html_file, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\n✅ HTML report generated: {html_file}")
    return html_file


def export_all_formats():
    """
    Export leads to all formats (CSV, Excel, HTML) and print summary.
    """
    print("\n" + "=" * 60)
    print("📁 EXPORTING LEADS")
    print("=" * 60)

    # Export CSV
    csv_file = export_leads_to_csv()

    # Print summary to console
    print_leads_summary()

    # Generate HTML report
    html_file = generate_html_report()

    # Try Excel export (if openpyxl is installed)
    excel_file = export_to_excel()

    print("\n" + "=" * 60)
    print("✅ EXPORT COMPLETE")
    print("=" * 60)
    if csv_file:
        print(f"📄 CSV: {csv_file}")
    if excel_file:
        print(f"📊 Excel: {excel_file}")
    if html_file:
        print(f"🌐 HTML: {html_file}")
    print("\n💡 Tip: Open the HTML file in your browser for a formatted view with images.")

    return csv_file


def export_simple():
    """
    Simple one-file export (no HTML, just CSV).
    """
    leads_file = "data/final_leads.csv"

    if not os.path.exists(leads_file):
        print("❌ No final_leads.csv found.")
        return None

    df = pd.read_csv(leads_file)
    if df.empty:
        print("❌ No leads found.")
        return None

    export_dir = "data/exports"
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_file = os.path.join(export_dir, f"leads_{timestamp}.csv")

    df.to_csv(export_file, index=False)
    print(f"\n✅ Leads exported to: {export_file}")
    print(f"   Total leads: {len(df)}")

    return export_file


if __name__ == "__main__":
    export_all_formats()